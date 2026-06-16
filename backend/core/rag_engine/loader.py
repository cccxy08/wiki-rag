"""文档加载器 — PDF/TXT/MD/DOCX/XLSX/XLS/CSV"""
from __future__ import annotations
from pathlib import Path


class LoaderMixin:
    def load_file(self, file_path: Path) -> list[str]:
        ext = file_path.suffix.lower()
        loader = self._loader_map.get(ext)
        if loader is None:
            raise ValueError(f"不支持的文件格式: {ext}（支持: {list(self._loader_map.keys())}")
        return loader(file_path)

    def _load_pdf(self, path: Path) -> list[str]:
        try:
            import fitz
            doc = fitz.open(str(path))
            texts = []
            for page in doc:
                t = page.get_text()
                if t.strip():
                    texts.append(t)
            return texts
        except ImportError:
            raise ImportError("请安装 pymupdf: pip install pymupdf")

    def _load_text(self, path: Path) -> list[str]:
        text = path.read_text(encoding="utf-8")
        return [text] if text.strip() else []

    def _load_markdown(self, path: Path) -> list[str]:
        text = path.read_text(encoding="utf-8")
        return [text] if text.strip() else []

    def _load_docx(self, path: Path) -> list[str]:
        try:
            from docx import Document
            doc = Document(str(path))
            return [p.text for p in doc.paragraphs if p.text.strip()]
        except ImportError:
            raise ImportError("请安装 python-docx: pip install python-docx")

    def _load_xlsx(self, path: Path) -> list[str]:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(min_row=1, values_only=True):
                    cleaned = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cleaned:
                        rows.append(" | ".join(cleaned))
                if rows:
                    texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            return texts
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

    def _load_xls(self, path: Path) -> list[str]:
        try:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            texts = []
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                rows = []
                for row_idx in range(ws.nrows):
                    cleaned = [str(ws.cell_value(row_idx, c)).strip()
                               for c in range(ws.ncols)
                               if str(ws.cell_value(row_idx, c)).strip()]
                    if cleaned:
                        rows.append(" | ".join(cleaned))
                if rows:
                    texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            return texts
        except ImportError:
            raise ImportError("请安装 xlrd: pip install xlrd")