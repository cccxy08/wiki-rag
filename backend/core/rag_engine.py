"""RAG 引擎 - 文档加载、切片、向量化、检索"""
import hashlib
import logging
from collections import defaultdict
from pathlib import Path
from typing import Optional, Union

from langchain_text_splitters import RecursiveCharacterTextSplitter
import chromadb
from chromadb.config import Settings as ChromaSettings

from core.config import settings
from core.llm_provider import get_llm
from core.retrieval import Reranker

logger = logging.getLogger(__name__)


# ==================== Embedding 抽象层 ====================

class BaseEmbedding:
    """Embedding 抽象基类，兼容 SentenceTransformer 的 encode / get_sentence_embedding_dimension 接口"""

    def encode(self, texts: Union[str, list[str]], show_progress_bar: bool = False, **kwargs) -> list[list[float]]:
        raise NotImplementedError

    def get_sentence_embedding_dimension(self) -> int:
        raise NotImplementedError


class LocalEmbedding(BaseEmbedding):
    """本地 SentenceTransformer Embedding（原方案，需 ~500MB 内存）"""

    def __init__(self, model_name: str, device: str = "cpu"):
        from sentence_transformers import SentenceTransformer
        self._model = SentenceTransformer(model_name, device=device)

    def encode(self, texts, show_progress_bar=False, **kwargs):
        result = self._model.encode(texts, show_progress_bar=show_progress_bar, **kwargs)
        return result

    def get_sentence_embedding_dimension(self) -> int:
        return self._model.get_sentence_embedding_dimension()


class ZhipuAPIEmbedding(BaseEmbedding):
    """智谱 Embedding API（零内存，推荐）"""

    # 智谱 embedding-3 固定 2048 维
    MODEL_DIMS = {
        "embedding-3": 2048,
        "embedding-2": 1024,
    }

    def __init__(self, model: str = "embedding-3", api_key: str = None, batch_size: int = 16):
        from openai import OpenAI
        self._model = model
        self._dimension = self.MODEL_DIMS.get(model, 2048)
        self._batch_size = batch_size
        api_key = api_key or settings.zhipu_api_key
        if not api_key:
            raise ValueError("智谱 Embedding 需要 ZHIPU_API_KEY，请在 .env 中配置")
        self._client = OpenAI(
            base_url="https://open.bigmodel.cn/api/paas/v4",
            api_key=api_key,
        )

    def encode(self, texts, show_progress_bar=False, **kwargs):
        import numpy as np
        was_single = isinstance(texts, str)
        if was_single:
            texts = [texts]
        all_embeddings = []
        for i in range(0, len(texts), self._batch_size):
            batch = texts[i:i + self._batch_size]
            try:
                response = self._client.embeddings.create(
                    model=self._model,
                    input=batch,
                )
                batch_embeddings = [item.embedding for item in response.data]
                all_embeddings.extend(batch_embeddings)
            except Exception as e:
                logger.error(f"智谱 Embedding API 调用失败 (batch {i // self._batch_size}): {e}")
                # 降级：返回零向量
                all_embeddings.extend([[0.0] * self._dimension] * len(batch))
        result = np.array(all_embeddings)
        # 单条查询时 squeeze 掉 batch 维度，与 sentence_transformers 行为一致
        if was_single:
            return result[0]
        return result

    def get_sentence_embedding_dimension(self) -> int:
        return self._dimension


class OpenAIAPIEmbedding(BaseEmbedding):
    """OpenAI Embedding API（零内存）"""

    # OpenAI 模型维度映射
    MODEL_DIMS = {
        "text-embedding-3-small": 1536,
        "text-embedding-3-large": 3072,
        "text-embedding-ada-002": 1536,
    }

    def __init__(self, model: str = "text-embedding-3-small", api_key: str = None,
                 base_url: str = None, batch_size: int = 16):
        from openai import OpenAI
        self._model = model
        self._dimension = self.MODEL_DIMS.get(model, 1536)
        self._batch_size = batch_size
        api_key = api_key or settings.openai_api_key
        base_url = base_url or settings.openai_base_url
        if not api_key:
            raise ValueError("OpenAI Embedding 需要 OPENAI_API_KEY，请在 .env 中配置")
        self._client = OpenAI(base_url=base_url, api_key=api_key)

    def encode(self, texts, show_progress_bar=False, **kwargs):
        import numpy as np
        was_single = isinstance(texts, str)
        if was_single:
            texts = [texts]
        all_embeddings = []
        for i in range(0, len(texts), self._batch_size):
            batch = texts[i:i + self._batch_size]
            try:
                response = self._client.embeddings.create(
                    model=self._model,
                    input=batch,
                )
                batch_embeddings = [item.embedding for item in response.data]
                all_embeddings.extend(batch_embeddings)
            except Exception as e:
                logger.error(f"OpenAI Embedding API 调用失败 (batch {i // self._batch_size}): {e}")
                all_embeddings.extend([[0.0] * self._dimension] * len(batch))
        result = np.array(all_embeddings)
        # 单条查询时 squeeze 掉 batch 维度，与 sentence_transformers 行为一致
        if was_single:
            return result[0]
        return result

    def get_sentence_embedding_dimension(self) -> int:
        return self._dimension


def create_embedding() -> BaseEmbedding:
    """工厂函数：根据 embedding_provider 配置创建 Embedding 实例"""
    provider = settings.embedding_provider
    logger.info(f"初始化 Embedding: provider={provider}")

    if provider == "zhipu":
        return ZhipuAPIEmbedding(
            model=settings.zhipu_embedding_model,
        )
    elif provider == "openai":
        return OpenAIAPIEmbedding(
            model=settings.openai_embedding_model,
        )
    else:  # local
        return LocalEmbedding(
            model_name=settings.embedding_model,
            device=settings.embedding_device,
        )


# ==================== RAG 引擎 ====================

class RAGEngine:
    """RAG 引擎 - 文档检索与问答"""

    _instance: Optional["RAGEngine"] = None

    @classmethod
    def get_instance(cls) -> "RAGEngine":
        """获取单例（避免多次加载 Embedding 模型占满内存）"""
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def __init__(self):
        self.llm = get_llm()
        self.embeddings = create_embedding()
        self.chroma_client = chromadb.PersistentClient(
            path=settings.chroma_persist_dir,
            settings=ChromaSettings(anonymized_telemetry=False),
        )
        self.collection = self.chroma_client.get_or_create_collection(
            name=settings.chroma_collection_name,
        )
        # Parent-Child: 父级 chunks 单独存一个 collection（仅 ID 查询，不参与向量检索）
        self.parent_collection = self.chroma_client.get_or_create_collection(
            name=settings.chroma_parent_collection_name,
        )
        self._loader_map = {
            ".pdf": self._load_pdf,
            ".txt": self._load_text,
            ".md": self._load_markdown,
            ".docx": self._load_docx,
            ".xlsx": self._load_xlsx,
            ".xls": self._load_xls,
            ".csv": self._load_text,  # CSV 按纯文本处理
        }
        # BM25 索引（懒加载）
        self._bm25 = None
        self._bm25_id_map = {}
        # Reranker（重排序）— 按配置决定是否加载
        if settings.reranker_enabled:
            self.reranker = Reranker(model_name_or_path=settings.reranker_model_path)
        else:
            self.reranker = None
            logger.info("Reranker 已禁用（reranker_enabled=False），节省 ~1.2GB 内存")

    # ========== 文档加载 ==========

    def load_file(self, file_path: Path) -> list[str]:
        """加载单个文件，返回文本段落列表"""
        ext = file_path.suffix.lower()
        loader = self._loader_map.get(ext)
        if loader is None:
            raise ValueError(f"不支持的文件格式: {ext}（支持: {list(self._loader_map.keys())}")
        return loader(file_path)

    def _load_pdf(self, path: Path) -> list[str]:
        try:
            import fitz  # pymupdf
            doc = fitz.open(str(path))
            texts = []
            for page in doc:
                t = page.get_text()
                if t.strip():
                    texts.append(t)
            return texts
        except ImportError:
            raise ImportError("请安装 pymupdf: pip install pymupdf")

    def _load_text(self, path: Path) -> list[str]:
        text = path.read_text(encoding="utf-8")
        return [text] if text.strip() else []

    def _load_markdown(self, path: Path) -> list[str]:
        text = path.read_text(encoding="utf-8")
        return [text] if text.strip() else []

    def _load_docx(self, path: Path) -> list[str]:
        try:
            from docx import Document
            doc = Document(str(path))
            return [p.text for p in doc.paragraphs if p.text.strip()]
        except ImportError:
            raise ImportError("请安装 python-docx: pip install python-docx")

    def _load_xlsx(self, path: Path) -> list[str]:
        """用 openpyxl 解析 .xlsx，每个 sheet 输出为文字表"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            texts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(min_row=1, values_only=True):
                    cleaned = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cleaned:
                        rows.append(" | ".join(cleaned))
                if rows:
                    texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            return texts
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

    def _load_xls(self, path: Path) -> list[str]:
        """用 xlrd 解析老版 .xls，每个 sheet 输出为文字表"""
        try:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            texts = []
            for sheet_name in wb.sheet_names():
                ws = wb.sheet_by_name(sheet_name)
                rows = []
                for row_idx in range(ws.nrows):
                    cleaned = [str(ws.cell_value(row_idx, c)).strip()
                               for c in range(ws.ncols)
                               if str(ws.cell_value(row_idx, c)).strip()]
                    if cleaned:
                        rows.append(" | ".join(cleaned))
                if rows:
                    texts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            return texts
        except ImportError:
            raise ImportError("请安装 xlrd: pip install xlrd")

    # ========== 切片 & 向量化 ==========

    def chunk_texts(self, texts: list[str], ext: str = "") -> list[str]:
        strategy = settings.chunk_strategy.get(ext.lstrip("."), {})
        cs = strategy.get("chunk_size", settings.chunk_size)
        co = strategy.get("chunk_overlap", settings.chunk_overlap)
        seps = strategy.get("separators", ["\n\n", "\n", ".", " ", ""])
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=cs,
            chunk_overlap=co,
            separators=seps,
        )
        full_text = "\n\n".join(texts)
        return splitter.split_text(full_text)

    def chunk_texts_parent_child(self, texts: list[str]) -> tuple[list[str], list[str], list[int]]:
        """
        Parent-Child 双层级切片。
        - child: 250 字小粒度，存 Chroma 做向量检索（精度高）
        - parent: 750 字大粒度，child → parent 映射后送给 LLM（上下文完整）
        返回: (child_chunks, parent_chunks, child_to_parent_map)
               child_to_parent_map[i] = parent 的下标
        """
        full_text = "\n\n".join(texts)

        # 1. Parent 层：750 字大切片（无重叠，作为上下文单元）
        parent_splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.parent_chunk_size,
            chunk_overlap=settings.parent_chunk_overlap,
            separators=["\n\n", "\n", ".", " ", ""],
        )
        parent_chunks = parent_splitter.split_text(full_text)
        if not parent_chunks:
            return [], [], []

        # 2. Child 层：每个 parent 内部切 250 字小子块
        child_splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.child_chunk_size,
            chunk_overlap=settings.child_chunk_overlap,
            separators=["\n\n", "\n", ".", " ", ""],
        )

        child_chunks = []
        child_to_parent = []
        for p_idx, parent in enumerate(parent_chunks):
            children = child_splitter.split_text(parent)
            child_chunks.extend(children)
            child_to_parent.extend([p_idx] * len(children))

        return child_chunks, parent_chunks, child_to_parent

    def index_document(self, file_path: Path) -> int:
        """索引一个文档：清理旧版本 → 加载 → Parent-Child 切片 → 向量化 → 存入 Chroma"""
        # P1.2: 增量索引 — 先清理同一文件的旧 chunks，防止重复堆积
        self._remove_document_chunks(file_path)

        try:
            texts = self.load_file(file_path)
        except Exception as e:
            raise RuntimeError(f"文档加载失败 [{file_path.name}]: {e}")

        if not texts:
            raise ValueError(f"文档内容为空: {file_path.name}")

        # ===== Parent-Child 双层级切片 =====
        child_chunks, parent_chunks, child_to_parent = self.chunk_texts_parent_child(texts)

        if not child_chunks:
            raise ValueError(f"切片后无内容: {file_path.name}")

        # ---- 1. 存入 Parent collection（仅 ID 查询用，不做向量检索） ----
        parent_ids = [f"p_{file_path.stem}_{i}" for i in range(len(parent_chunks))]
        parent_metadatas = [
            {"source": file_path.name, "parent_index": i, "path": str(file_path)}
            for i in range(len(parent_chunks))
        ]
        # Parent 不需要有效 embedding，填零向量即可（永远不会被向量搜索）
        dim = self.embeddings.get_sentence_embedding_dimension()
        parent_embeddings = [[0.0] * dim for _ in parent_chunks]
        self.parent_collection.add(
            ids=parent_ids,
            embeddings=parent_embeddings,
            documents=parent_chunks,
            metadatas=parent_metadatas,
        )

        # ---- 2. 存入 Child collection（参与向量 + BM25 混合检索） ----
        child_ids = []
        for i, chunk in enumerate(child_chunks):
            chunk_hash = hashlib.md5(chunk.encode()).hexdigest()[:12]
            child_ids.append(f"{file_path.stem}_{i}_{chunk_hash}")

        embeddings = self.embeddings.encode(child_chunks, show_progress_bar=False).tolist()
        metadatas = [
            {
                "source": file_path.name,
                "chunk_index": i,
                "path": str(file_path),
                "parent_id": f"p_{file_path.stem}_{child_to_parent[i]}",
            }
            for i in range(len(child_chunks))
        ]

        self.collection.add(
            ids=child_ids,
            embeddings=embeddings,
            documents=child_chunks,
            metadatas=metadatas,
        )

        # 重建 BM25 索引（标记为脏，下次检索时懒加载）
        self._bm25 = None

        return len(child_chunks)

    def _remove_document_chunks(self, file_path: Path) -> int:
        """安全删除同一文件的旧 chunks（按完整路径精确匹配，不会误删其他文档）"""
        deleted = 0
        try:
            existing = self.collection.get(where={"path": str(file_path)})
            old_ids = existing.get("ids", [])
            if old_ids:
                self.collection.delete(ids=old_ids)
                deleted = len(old_ids)
        except Exception:
            # Chroma 首次创建 collection 时 where 查询可能不稳定，降级跳过
            pass

        # Parent-Child: 同步清理 parent collection
        try:
            existing_parents = self.parent_collection.get(where={"path": str(file_path)})
            old_parent_ids = existing_parents.get("ids", [])
            if old_parent_ids:
                self.parent_collection.delete(ids=old_parent_ids)
        except Exception:
            pass

        return deleted

    # ========== 检索 ==========

    def retrieve(self, query: str, top_k: int = None) -> list[dict]:
        """默认检索：向量 + BM25 混合检索"""
        return self.retrieve_with_bm25(query, top_k)

    def retrieve_vector_only(self, query: str, top_k: int = None) -> list[dict]:
        """纯向量检索（用于对比/调试）"""
        if top_k is None:
            top_k = settings.retrieval_top_k

        query_embedding = self.embeddings.encode(query, show_progress_bar=False).tolist()

        # P1.1: 取 top_k * 3 个 chunk，去重后再截断，确保 top_k 个来自不同文档
        fetch_k = top_k * 3
        results = self.collection.query(
            query_embeddings=[query_embedding],
            n_results=fetch_k,
        )

        documents = []
        if results["ids"] and results["ids"][0]:
            for i, doc_id in enumerate(results["ids"][0]):
                score = results["distances"][0][i] if results.get("distances") else 0
                # 相似度阈值过滤
                # ChromaDB 默认用 L2 距离：越小越相似，低于阈值的才保留（距离过大的不相似）
                if score > settings.similarity_threshold:
                    continue
                documents.append({
                    "id": doc_id,
                    "content": results["documents"][0][i] if results.get("documents") else "",
                    "metadata": results["metadatas"][0][i] if results.get("metadatas") else {},
                    "score": score,
                })

        # 按 source 去重：同一文档只保留最高分的 chunk
        docs = self._dedup_by_source(documents, top_k)

        # Parent-Child 映射
        if settings.parent_child_enabled:
            docs = self._map_children_to_parents(docs)

        # Reranker 重排序（仅在启用时）
        if self.reranker and self.reranker.available and len(docs) > 1:
            docs = self.reranker.rerank(query, docs, top_k=settings.rerank_top_k)

        return docs

    def _dedup_by_source(self, docs: list[dict], top_k: int) -> list[dict]:
        """按 source 分组去重，每组只保留距离最近的 chunk，返回 top_k 个不同来源"""
        best = {}
        for d in docs:
            src = d.get("metadata", {}).get("source", "__UNKNOWN__")
            # ChromaDB 距离：越小越相关，所以保留分数最低的 chunk
            if src not in best or d["score"] < best[src]["score"]:
                best[src] = d
        # 距离升序：最近（最相关）的排最前
        return sorted(best.values(), key=lambda x: x["score"])[:top_k]

    # ========== BM25 混合检索 ==========

    def _tokenize(self, text: str) -> list[str]:
        """对中文文本做 jieba 分词，英文/数字保留原词"""
        import jieba
        import re
        # 先用正则提取中文段落和英数段落
        tokens = []
        for segment in re.findall(r'[\u4e00-\u9fff]+|[a-zA-Z0-9]+', text):
            if re.match(r'[\u4e00-\u9fff]+', segment):
                tokens.extend(jieba.lcut(segment))  # jieba 分词
            else:
                tokens.append(segment.lower())      # 英数转小写
        return tokens

    def _rebuild_bm25_index(self):
        """从 ChromaDB 读取所有文档，重建 BM25 索引"""
        from rank_bm25 import BM25Okapi
        all_data = self.collection.get(include=["documents"])
        if not all_data or not all_data.get("documents"):
            self._bm25 = None
            self._bm25_id_map = {}
            return
        tokenized_corpus = []
        self._bm25_id_map = {}
        for i, doc_text in enumerate(all_data["documents"]):
            tokenized_corpus.append(self._tokenize(doc_text))
            self._bm25_id_map[all_data["ids"][i]] = i  # doc_id → corpus index
        self._bm25 = BM25Okapi(tokenized_corpus)

    def retrieve_with_bm25(self, query: str, top_k: int = None) -> list[dict]:
        """混合检索：向量 + BM25 加权融合"""
        if top_k is None:
            top_k = settings.retrieval_top_k

        # ---- 1. 向量检索（原样取，不做阈值过滤，留给融合后统一去重） ----
        query_embedding = self.embeddings.encode(query, show_progress_bar=False).tolist()
        fetch_k = top_k * 5
        results = self.collection.query(
            query_embeddings=[query_embedding],
            n_results=fetch_k,
        )

        vector_docs = []
        if results["ids"] and results["ids"][0]:
            for i, doc_id in enumerate(results["ids"][0]):
                score = results["distances"][0][i] if results.get("distances") else 0
                vector_docs.append({
                    "id": doc_id,
                    "content": results["documents"][0][i] if results.get("documents") else "",
                    "metadata": results["metadatas"][0][i] if results.get("metadatas") else {},
                    "vector_l2": score,
                })

        if not vector_docs:
            return []

        # ---- 2. BM25 检索 ----
        if self._bm25 is None:
            self._rebuild_bm25_index()

        query_tokens = self._tokenize(query)
        bm25_scores_all = self._bm25.get_scores(query_tokens) if self._bm25 else []

        # 为每个 vector_doc 匹配 BM25 分
        for d in vector_docs:
            idx = self._bm25_id_map.get(d["id"])
            d["bm25_raw"] = bm25_scores_all[idx] if idx is not None and len(bm25_scores_all) > 0 else 0.0

        # ---- 3. 归一化 ----
        # 向量 L2 距离 → 相似度 (0~1)：sim = 1 / (1 + l2)
        # BM25 raw → [0,1]：min-max 归一化
        l2_vals = [d["vector_l2"] for d in vector_docs]
        bm25_vals = [d["bm25_raw"] for d in vector_docs]
        max_l2 = max(l2_vals) if l2_vals else 1.0
        max_bm25 = max(bm25_vals) if bm25_vals else 1.0

        for d in vector_docs:
            d["vector_sim"] = 1.0 / (1.0 + d["vector_l2"])
            d["bm25_norm"] = d["bm25_raw"] / max_bm25 if max_bm25 > 0 else 0.0

        # ---- 4. 加权融合 ----
        bw = settings.bm25_weight
        for d in vector_docs:
            d["hybrid_score"] = (1 - bw) * d["vector_sim"] + bw * d["bm25_norm"]

        # ---- 5. 按 hybrid_score 降序排列 + 来源去重 ----
        vector_docs.sort(key=lambda x: x["hybrid_score"], reverse=True)

        # 去重：同 source 最多保留 hybrid_score 最高的 2 个 chunk
        # （避免暴力 1chunk/文件 导致同一文件的低频子话题被丢弃）
        buckets = defaultdict(list)
        for d in vector_docs:
            src = d.get("metadata", {}).get("source", "__UNKNOWN__")
            buckets[src].append(d)
        best = []
        for src, items in buckets.items():
            items.sort(key=lambda x: x["hybrid_score"], reverse=True)
            best.extend(items[:2])

        merged = sorted(best, key=lambda x: x["hybrid_score"], reverse=True)[:top_k]

        # 把 hybrid_score 映射回统一的 score 字段
        for d in merged:
            d["score"] = d["hybrid_score"]

        # ===== Parent-Child 映射：child → parent（去重合并） =====
        if settings.parent_child_enabled:
            merged = self._map_children_to_parents(merged)

        # ===== Reranker 重排序（仅在启用时） =====
        if self.reranker and self.reranker.available and len(merged) > 1:
            merged = self.reranker.rerank(query, merged, top_k=settings.rerank_top_k)

        return merged

    def _map_children_to_parents(self, child_docs: list[dict]) -> list[dict]:
        """将检索到的 child chunks 映射到 parent chunks，去重合并后返回"""
        parent_ids_seen = set()
        parent_docs = []
        for child in child_docs:
            pid = child.get("metadata", {}).get("parent_id")
            if not pid or pid in parent_ids_seen:
                continue
            parent_ids_seen.add(pid)

            try:
                parent_data = self.parent_collection.get(ids=[pid])
                if parent_data and parent_data.get("documents"):
                    parent_docs.append({
                        "id": pid,
                        "content": parent_data["documents"][0],
                        "metadata": parent_data["metadatas"][0] if parent_data.get("metadatas") else {},
                        "score": child["score"],
                    })
            except Exception:
                continue
        return parent_docs

    # ========== 问答 ==========

    def answer(self, question: str, context_docs: list[dict]) -> str:
        """基于检索结果生成回答"""
        if not context_docs:
            return "未找到相关文档，无法回答此问题。"

        context_text = "\n\n---\n\n".join([
            f"[来源: {doc['metadata'].get('source', 'unknown')}]\n{doc['content']}"
            for doc in context_docs[:settings.rerank_top_k]
        ])

        prompt = f"""你是一个企业知识库助手。请根据以下检索到的文档内容回答用户问题。

## 输出格式要求（严格遵守）

回答必须是 **纯 HTML 片段**（不含 html/body/head 标签），渲染美观，结构清晰。

### HTML 结构模板

```html
<div class="rag-answer">
  <div class="rag-section">
    <span class="rag-label">结论</span>
    <span class="rag-content"><strong>结论内容加粗</strong></span>
  </div>
  <div class="rag-section">
    <span class="rag-label">依据</span>
    <ul class="rag-list">
      <li><span class="rag-source">文档名称</span>：原文摘要</li>
    </ul>
  </div>
  <div class="rag-section">
    <span class="rag-label">补充</span>
    <span class="rag-content">额外上下文</span>
  </div>
  <div class="rag-section rag-footer">
    <span class="rag-label">来源</span>
    <span class="rag-content">文件1、文件2</span>
  </div>
</div>
```

### 规则
1. 只依据提供的文档内容回答，不要编造
2. **语义匹配**：用户用词可能不是文档中的精确用词（如"老板"可能对应文档中的"董事长/创始人/CEO"），请根据语义理解匹配，不要因为用词不同就判定"未找到"
3. 如果确实找不到相关信息，返回：<div class="rag-answer">未在文档中找到相关信息。</div>（仅在检索到的所有文档确实都不包含任何相关内容时才返回）
4. 关键数字/名称用 <strong> 加粗
5. 每个依据项标注来源文档名称
6. <span class="rag-source"> 包裹文档名，冒号后跟原文摘要
7. 不要输出 ```html 代码块包裹，直接输出 HTML 片段
8. **来源权威性**：当多个来源信息冲突时，按以下优先级取信：配置文件/制度文档 > 官方公告/通知 > 会议纪要 > 其他
9. **数值计算**：涉及数值比较或计算时，必须列出计算过程和算式，不要心算

---

检索到的文档：
{context_text}

用户问题：{question}

请按上述格式用中文回答：
"""
        return self.llm.chat([
            {"role": "system", "content": "你是企业知识库助手，只依据检索到的文档回答。注意语义匹配：用户用词可能不等于文档中的精确用词（如'老板'可能对应'董事长/CEO/创始人'），允许合理推断。回答必须是纯HTML片段，不含html/body/head标签，使用rag-answer/rag-section/rag-label/rag-content/rag-list/rag-source等class名。重要规则：来源冲突时配置文件>公告>会议纪要；数值比较必须列出计算过程。"},
            {"role": "user", "content": prompt}
        ], label="answer")

    # ========== 工具方法 ==========

    def collection_count(self) -> int:
        """向量库中的文档数量"""
        return self.collection.count()

    def clear_collection(self):
        """清空向量库（child + parent）"""
        try:
            self.chroma_client.delete_collection(settings.chroma_collection_name)
        except Exception:
            pass
        try:
            self.chroma_client.delete_collection(settings.chroma_parent_collection_name)
        except Exception:
            pass
        self.collection = self.chroma_client.get_or_create_collection(
            name=settings.chroma_collection_name,
        )
        self.parent_collection = self.chroma_client.get_or_create_collection(
            name=settings.chroma_parent_collection_name,
        )
